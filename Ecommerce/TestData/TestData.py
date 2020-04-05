import openpyxl


class TestData:
    HomePageData = [{"Name": "Sidheshwar", "Email": "Sid@test.com", "Gender": "Male"},
                    {"Name": "Manvi", "Email": "Tondare@gmail.com", "Gender": "Female"}]

    @staticmethod
    def getTestData():

        workbook = openpyxl.load_workbook(
            "C:\\Users\\Sidheshwar.Tondare\\PycharmProjects\\Ecommerce\\TestData\\EcommerceData.xlsx")
        sheet = workbook.active
        data = {}
        listData = []
        for i in range(1, sheet.max_row + 1):
            if i > 1:
                for j in range(2, sheet.max_column + 1):
                    data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                listData.append(data.copy())

        return listData
