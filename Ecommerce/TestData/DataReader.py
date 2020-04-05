import openpyxl

workbook = openpyxl.load_workbook(
    "C:\\Users\\Sidheshwar.Tondare\\PycharmProjects\\Ecommerce\\TestData\\EcommerceData.xlsx")
sheet = workbook.active
listData = []
data = {}
cell = sheet.cell(row=1, column=1)
print(cell.value)
print(sheet.cell(row=2, column=2).value)
# sheet.cell(row=2,column=2).value="manavi"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_column)
print(sheet.max_row)
for i in range(1, sheet.max_row + 1):
    if i > 1:
        for j in range(2, sheet.max_column + 1):
            data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        listData.append(data.copy())
print(listData)
